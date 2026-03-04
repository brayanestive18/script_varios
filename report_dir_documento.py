#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Genera un reporte Excel de asistencia DIR a partir de un CSV con documentos.

Uso:
    python report_dir_documento.py alumnos.csv
"""

import sys
import os
import csv
from datetime import datetime

import mysql.connector
from mysql.connector import Error
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

load_dotenv()

DB_CONFIG = {
    'host':     os.getenv('DB_HOST', 'localhost'),
    'user':     os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', ''),
    'database': os.getenv('DB_NAME', 'erp'),
}

COLUMNAS = ['ID Alumno', 'Nombre', 'Grupo DIR', '% Asistencia', 'Clases Asistidas', 'Clases Totales']


def leer_documentos(csv_path: str) -> list[str]:
    """Lee la primera columna del CSV y devuelve los valores únicos como strings."""
    if not os.path.exists(csv_path):
        print(f"Error: no se encontró el archivo '{csv_path}'")
        sys.exit(1)

    documentos = []
    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader, None)
        for row in reader:
            if row and row[0].strip():
                documentos.append(row[0].strip())

    if not documentos:
        print("Error: el CSV no contiene documentos.")
        sys.exit(1)

    # Deduplicar preservando orden
    vistos = set()
    duplicados = []
    unicos = []
    for d in documentos:
        if d in vistos:
            duplicados.append(d)
        else:
            vistos.add(d)
            unicos.append(d)

    if header:
        print(f"  Columna usada: '{header[0]}'")
    print(f"  {len(documentos)} filas leídas → {len(unicos)} únicos")
    if duplicados:
        print(f"  {len(duplicados)} duplicados eliminados: {', '.join(duplicados)}")
    return unicos


def consultar(documentos: list[str]) -> list[tuple]:
    """Ejecuta la consulta SQL y devuelve los resultados."""
    placeholders = ', '.join(['%s'] * len(documentos))

    sql = f"""
WITH
alumnos_input AS (
  SELECT al.id AS id_alumno,
    TRIM(CONCAT(u.nombre1,' ',COALESCE(u.nombre2,''),' ',u.apellido1,' ',COALESCE(u.apellido2,''))) AS nombre
  FROM alumno al
  JOIN usuario u ON al.id = u.id AND al.dni = u.dni
  WHERE al.id IN ({placeholders})
),
grupos_dir AS (
  SELECT g.id AS grupo_id
  FROM grupo g
  WHERE g.materia IN (34, 48, 55, 65, 68, 69, 74)
    AND g.id NOT IN (1174, 294, 516, 647)
),
tc AS (
  SELECT c.grupo, COUNT(*) AS total_clases
  FROM clase c
  JOIN grupos_dir gd ON gd.grupo_id = c.grupo
  WHERE c.est_clase = 2 AND c.url_video IS NOT NULL
  GROUP BY c.grupo
),
ac AS (
  SELECT a.grupo, a.id_alumno, COUNT(DISTINCT a.clase) AS clases_con_asistencia
  FROM asistencia_clase a
  JOIN clase c
    ON c.grupo = a.grupo
   AND c.id = a.clase
   AND c.est_clase = 2
   AND c.url_video IS NOT NULL
  JOIN grupos_dir gd ON gd.grupo_id = a.grupo
  GROUP BY a.grupo, a.id_alumno
),
dir_mat AS (
  SELECT m.id_alumno, m.grupo
  FROM matricula_materia m
  JOIN grupos_dir gd ON gd.grupo_id = m.grupo
),
per_alumno_grupo AS (
  SELECT
    ai.id_alumno,
    ai.nombre,
    COALESCE(CONCAT('Grupo ', dm.grupo), 'No matriculado en DIR') AS grupo,
    COALESCE(tc.total_clases, 0)          AS clases_totales,
    COALESCE(ac.clases_con_asistencia, 0) AS clases_con_asistencia,
    ROUND(
      CASE WHEN COALESCE(tc.total_clases, 0) > 0
           THEN (COALESCE(ac.clases_con_asistencia, 0) * 100.0) / tc.total_clases
           ELSE 0
      END, 2
    ) AS porcentaje
  FROM alumnos_input ai
  LEFT JOIN dir_mat dm        ON dm.id_alumno = ai.id_alumno
  LEFT JOIN tc                ON tc.grupo = dm.grupo
  LEFT JOIN ac                ON ac.grupo = dm.grupo AND ac.id_alumno = ai.id_alumno
  GROUP BY ai.id_alumno, ai.nombre, dm.grupo, tc.total_clases, ac.clases_con_asistencia
)
SELECT id_alumno, nombre, grupo, porcentaje, clases_con_asistencia, clases_totales
FROM (
  SELECT *, ROW_NUMBER() OVER (PARTITION BY id_alumno ORDER BY porcentaje DESC, clases_con_asistencia DESC) AS rn
  FROM per_alumno_grupo
) t
WHERE rn = 1
ORDER BY nombre
"""

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute(sql, documentos)
        filas = cursor.fetchall()
        cursor.close()
        conn.close()
        return filas
    except Error as e:
        print(f"Error de base de datos: {e}")
        sys.exit(1)


def exportar_excel(filas: list[tuple], ruta_excel: str):
    """Escribe los resultados en un Excel con tabla formateada."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte DIR"

    # Encabezados
    ws.append(COLUMNAS)

    # Datos
    for fila in filas:
        ws.append(list(fila))

    # Tabla Excel
    ultima_fila = len(filas) + 1
    ultima_col = get_column_letter(len(COLUMNAS))
    ref = f"A1:{ultima_col}{ultima_fila}"

    tabla = Table(displayName="ReporteDir", ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)

    # Ancho de columnas automático
    anchos = [len(col) for col in COLUMNAS]
    for fila in filas:
        for i, celda in enumerate(fila):
            anchos[i] = max(anchos[i], len(str(celda)))

    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho + 4

    # Columna % centrada
    col_pct = 4
    for fila_idx in range(2, ultima_fila + 1):
        ws.cell(row=fila_idx, column=col_pct).alignment = Alignment(horizontal='center')

    wb.save(ruta_excel)


def main():
    if len(sys.argv) < 2:
        print(f"Uso: python {os.path.basename(__file__)} <archivo.csv>")
        sys.exit(1)

    csv_path = sys.argv[1]
    nombre_base = sys.argv[2] if len(sys.argv) >= 3 else None

    if not nombre_base:
        nombre_base = input("Nombre del archivo Excel de salida: ").strip()
    if not nombre_base:
        print("Error: el nombre no puede estar vacío.")
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y%m%d%H%M")
    nombre_excel = f"{nombre_base}_Report_Dir_{timestamp}.xlsx"

    print(f"\nLeyendo CSV...")
    documentos = leer_documentos(csv_path)

    print("Consultando base de datos...")
    filas = consultar(documentos)
    en_dir = sum(1 for f in filas if f[2] != 'No matriculado en DIR')
    sin_dir = len(filas) - en_dir
    print(f"  {len(filas)} alumnos encontrados  ({en_dir} en DIR, {sin_dir} no matriculados en DIR)")

    print(f"Generando Excel '{nombre_excel}'...")
    exportar_excel(filas, nombre_excel)
    print(f"  Listo -> {os.path.abspath(nombre_excel)}")


if __name__ == "__main__":
    main()
